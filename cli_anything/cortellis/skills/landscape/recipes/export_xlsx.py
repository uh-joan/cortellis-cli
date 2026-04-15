#!/usr/bin/env python3
"""
export_xlsx.py — Export landscape analysis as Excel workbook.

Reads scored CSVs from a landscape directory and produces a multi-sheet
Excel workbook with auto-filters and conditional formatting.

Usage: python3 export_xlsx.py <landscape_dir> [indication_name] [--output PATH]
"""

import os
import sys

# Allow running as standalone script
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "..", "..", ".."))

from cli_anything.cortellis.utils.data_helpers import read_csv_safe, safe_float, safe_int
from cli_anything.cortellis.utils.wiki import slugify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter



# ---------------------------------------------------------------------------
# Workbook creation and styling
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def create_workbook() -> Workbook:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    return wb


def style_header_row(ws, num_cols: int) -> None:
    """Apply navy fill + white bold text to the first row."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_size_columns(ws, headers: list) -> None:
    """Set column widths based on header length."""
    for col_idx, header in enumerate(headers, 1):
        width = max(len(str(header)) + 4, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def finalize_sheet(ws, headers: list) -> None:
    """Apply auto-filter, freeze panes, and auto-size columns."""
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"
    ws.freeze_panes = "A2"
    auto_size_columns(ws, headers)


# ---------------------------------------------------------------------------
# Sheet: Pipeline by Phase
# ---------------------------------------------------------------------------

PHASE_FILES = ["launched", "phase3", "phase2", "phase1", "discovery", "other"]

PIPELINE_HEADERS = ["Drug Name", "ID", "Phase", "Indication", "Mechanism", "Company"]


def add_pipeline_sheet(wb: Workbook, landscape_dir: str) -> None:
    ws = wb.create_sheet("Pipeline by Phase")

    ws.append(PIPELINE_HEADERS)
    style_header_row(ws, len(PIPELINE_HEADERS))

    for phase in PHASE_FILES:
        rows = read_csv_safe(os.path.join(landscape_dir, f"{phase}.csv"))
        for row in rows:
            drug_name = row.get("name") or row.get("drug_name") or row.get("drug") or ""
            drug_id = row.get("id") or row.get("drug_id") or ""
            phase_val = row.get("phase") or row.get("development_phase") or phase
            indication = row.get("indication") or row.get("indication_name") or ""
            mechanism = row.get("mechanism") or row.get("moa") or row.get("mechanism_of_action") or ""
            company = row.get("company") or row.get("company_name") or ""
            ws.append([drug_name, drug_id, phase_val, indication, mechanism, company])

    finalize_sheet(ws, PIPELINE_HEADERS)

    # Apply body font to data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT


# ---------------------------------------------------------------------------
# Sheet: Company Rankings
# ---------------------------------------------------------------------------

COMPANY_HEADERS = [
    "Rank", "Company", "Tier", "CPI Score", "Position",
    "Pipeline Breadth", "Phase Score", "Mechanism Diversity",
    "Deal Activity", "Trial Intensity",
]


def add_company_sheet(wb: Workbook, landscape_dir: str) -> None:
    ws = wb.create_sheet("Company Rankings")

    ws.append(COMPANY_HEADERS)
    style_header_row(ws, len(COMPANY_HEADERS))

    scores = read_csv_safe(os.path.join(landscape_dir, "strategic_scores.csv"))
    for rank, row in enumerate(scores, 1):
        cpi = safe_float(row.get("cpi_score"))
        ws.append([
            rank,
            row.get("company", ""),
            row.get("cpi_tier", ""),
            cpi,
            row.get("position", ""),
            safe_int(row.get("pipeline_breadth")),
            safe_float(row.get("phase_score")),
            safe_int(row.get("mechanism_diversity")),
            safe_int(row.get("deal_activity")),
            safe_int(row.get("trial_intensity")),
        ])

    finalize_sheet(ws, COMPANY_HEADERS)

    # Apply body font to data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT

    # Conditional formatting on CPI Score (column D = index 4)
    if len(scores) > 0:
        last_row = len(scores) + 1
        cpi_range = f"D2:D{last_row}"

        # Green: CPI >= 60
        ws.conditional_formatting.add(
            cpi_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["60"], fill=GREEN_FILL),
        )
        # Yellow: 30 <= CPI < 60
        ws.conditional_formatting.add(
            cpi_range,
            CellIsRule(operator="between", formula=["30", "59.999"], fill=YELLOW_FILL),
        )
        # Red: CPI < 30
        ws.conditional_formatting.add(
            cpi_range,
            CellIsRule(operator="lessThan", formula=["30"], fill=RED_FILL),
        )


# ---------------------------------------------------------------------------
# Sheet: Mechanism Analysis
# ---------------------------------------------------------------------------

MECHANISM_HEADERS = [
    "Mechanism", "Active Count", "Launched", "P3", "P2", "P1",
    "Discovery", "Companies", "Crowding Index",
]


def add_mechanism_sheet(wb: Workbook, landscape_dir: str) -> None:
    ws = wb.create_sheet("Mechanism Analysis")

    ws.append(MECHANISM_HEADERS)
    style_header_row(ws, len(MECHANISM_HEADERS))

    rows = read_csv_safe(os.path.join(landscape_dir, "mechanism_scores.csv"))
    for row in rows:
        ws.append([
            row.get("mechanism", ""),
            safe_int(row.get("active_count")),
            safe_int(row.get("launched")),
            safe_int(row.get("phase3")),
            safe_int(row.get("phase2")),
            safe_int(row.get("phase1")),
            safe_int(row.get("discovery")),
            safe_int(row.get("company_count")),
            safe_int(row.get("crowding_index")),
        ])

    finalize_sheet(ws, MECHANISM_HEADERS)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT


# ---------------------------------------------------------------------------
# Sheet: Deals
# ---------------------------------------------------------------------------

DEALS_BASE_HEADERS = ["Title", "ID", "Principal", "Partner", "Type", "Date"]
DEALS_FINANCIAL_HEADERS = ["Upfront", "Milestones", "Royalties", "Total Value"]


def add_deals_sheet(wb: Workbook, landscape_dir: str) -> None:
    ws = wb.create_sheet("Deals")

    deals = read_csv_safe(os.path.join(landscape_dir, "deals.csv"))

    # Optional left-join with deal_financials.csv on deal_id
    financials_path = os.path.join(landscape_dir, "deal_financials.csv")
    financials_map = {}
    has_financials = os.path.exists(financials_path)
    if has_financials:
        fin_rows = read_csv_safe(financials_path)
        for fin in fin_rows:
            deal_id = fin.get("deal_id") or fin.get("id") or ""
            if deal_id:
                financials_map[deal_id] = fin

    headers = DEALS_BASE_HEADERS + (DEALS_FINANCIAL_HEADERS if has_financials else [])
    ws.append(headers)
    style_header_row(ws, len(headers))

    for row in deals:
        deal_id = row.get("id") or row.get("deal_id") or ""
        base = [
            row.get("title") or row.get("deal_title") or "",
            deal_id,
            row.get("principal") or "",
            row.get("partner") or "",
            row.get("type") or row.get("deal_type") or "",
            row.get("date") or row.get("deal_date") or "",
        ]
        if has_financials:
            fin = financials_map.get(deal_id, {})
            base += [
                fin.get("upfront_payment") or fin.get("upfront") or "",
                fin.get("milestone_payments") or fin.get("milestones") or "",
                fin.get("royalty_rate") or fin.get("royalties") or "",
                fin.get("total_deal_value") or fin.get("total_value") or "",
            ]
        ws.append(base)

    finalize_sheet(ws, headers)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT


# ---------------------------------------------------------------------------
# Sheet: Opportunity Matrix
# ---------------------------------------------------------------------------

OPPORTUNITY_HEADERS = [
    "Mechanism", "Status", "Total", "Launched", "P3", "P2", "P1",
    "Discovery", "Companies", "Opportunity Score",
]


def add_opportunity_sheet(wb: Workbook, landscape_dir: str) -> None:
    ws = wb.create_sheet("Opportunity Matrix")

    ws.append(OPPORTUNITY_HEADERS)
    style_header_row(ws, len(OPPORTUNITY_HEADERS))

    rows = read_csv_safe(os.path.join(landscape_dir, "opportunity_matrix.csv"))
    for row in rows:
        ws.append([
            row.get("mechanism", ""),
            row.get("status", ""),
            safe_int(row.get("total")),
            safe_int(row.get("launched")),
            safe_int(row.get("phase3")),
            safe_int(row.get("phase2")),
            safe_int(row.get("phase1")),
            safe_int(row.get("discovery")),
            safe_int(row.get("companies")),
            safe_float(row.get("opportunity_score")),
        ])

    finalize_sheet(ws, OPPORTUNITY_HEADERS)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def export_xlsx(landscape_dir: str, indication_name: str, output_path: str = None) -> str:
    """Build the Excel workbook and write it to output_path.

    Returns the path to the written file.
    """
    if not output_path:
        slug = slugify(indication_name) if indication_name else slugify(
            os.path.basename(os.path.abspath(landscape_dir))
        )
        output_path = os.path.join(landscape_dir, f"{slug}-landscape.xlsx")

    wb = create_workbook()

    add_pipeline_sheet(wb, landscape_dir)
    add_company_sheet(wb, landscape_dir)
    add_mechanism_sheet(wb, landscape_dir)
    add_deals_sheet(wb, landscape_dir)
    add_opportunity_sheet(wb, landscape_dir)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print(
            "Usage: export_xlsx.py <landscape_dir> [indication_name] [--output PATH]",
            file=sys.stderr,
        )
        sys.exit(1)

    landscape_dir = sys.argv[1]
    indication_name = (
        sys.argv[2] if len(sys.argv) > 2 and not sys.argv[2].startswith("--") else None
    )

    # Parse --output
    output_path = None
    for i, arg in enumerate(sys.argv):
        if arg == "--output" and i + 1 < len(sys.argv):
            output_path = sys.argv[i + 1]

    if not os.path.isdir(landscape_dir):
        print(f"Error: landscape directory not found: {landscape_dir}", file=sys.stderr)
        sys.exit(1)

    if not indication_name:
        indication_name = os.path.basename(landscape_dir).replace("-", " ").title()

    out = export_xlsx(landscape_dir, indication_name, output_path)
    print(f"Exported: {out}")


if __name__ == "__main__":
    main()
