"""Tests for PPTX and XLSX export recipes."""

import csv
import os
import sys

import pytest


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_csv(path, header, rows):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for row in rows:
            w.writerow(row)


def _create_mock_landscape(tmp_path):
    """Create a minimal but valid landscape directory for export testing."""
    d = tmp_path / "raw" / "test-indication"
    d.mkdir(parents=True)
    sd = str(d)

    # Phase CSVs
    header = ["name", "id", "phase", "indication", "mechanism", "company", "source"]
    _write_csv(os.path.join(sd, "launched.csv"), header, [
        ["Drug A", "101", "Launched", "Test", "Mechanism X", "Company A", "src"],
        ["Drug B", "102", "Launched", "Test", "Mechanism Y", "Company B", "src"],
    ])
    _write_csv(os.path.join(sd, "phase3.csv"), header, [
        ["Drug C", "103", "Phase 3", "Test", "Mechanism X", "Company C", "src"],
    ])
    _write_csv(os.path.join(sd, "phase2.csv"), header, [
        ["Drug D", "104", "Phase 2", "Test", "Mechanism Z", "Company A", "src"],
        ["Drug E", "105", "Phase 2", "Test", "Mechanism X", "Company D", "src"],
    ])
    _write_csv(os.path.join(sd, "phase1.csv"), header, [])
    _write_csv(os.path.join(sd, "discovery.csv"), header, [])
    _write_csv(os.path.join(sd, "other.csv"), header, [])

    # Strategic scores
    _write_csv(os.path.join(sd, "strategic_scores.csv"),
        ["company", "cpi_tier", "cpi_score", "pipeline_breadth", "phase_score",
         "mechanism_diversity", "deal_activity", "trial_intensity", "position"],
        [
            ["Company A", "A", "85.0", "5", "40.0", "3", "4", "6", "Leader"],
            ["Company B", "B", "55.0", "3", "25.0", "2", "2", "3", "Challenger"],
            ["Company C", "C", "30.0", "2", "15.0", "1", "1", "1", "Emerging"],
        ])

    # Mechanism scores
    _write_csv(os.path.join(sd, "mechanism_scores.csv"),
        ["mechanism", "active_count", "launched", "phase3", "phase2", "phase1",
         "discovery", "company_count", "crowding_index"],
        [
            ["Mechanism X", "4", "1", "1", "1", "0", "0", "3", "12"],
            ["Mechanism Y", "1", "1", "0", "0", "0", "0", "1", "1"],
        ])

    # Opportunity matrix
    _write_csv(os.path.join(sd, "opportunity_matrix.csv"),
        ["mechanism", "launched", "phase3", "phase2", "phase1", "discovery",
         "total", "companies", "status", "opportunity_score"],
        [
            ["Mechanism X", "1", "1", "1", "0", "0", "3", "3", "Active", "0.5"],
            ["Mechanism Z", "0", "0", "1", "0", "0", "1", "1", "Emerging", "0.8"],
        ])

    # Deals
    _write_csv(os.path.join(sd, "deals.csv"),
        ["title", "id", "principal", "partner", "type", "date"],
        [
            ["Deal 1", "D001", "Company A", "Company B", "Licensing", "2026-03-01"],
            ["Deal 2", "D002", "Company C", "Company A", "Co-development", "2026-02-15"],
        ])

    return sd


# ---------------------------------------------------------------------------
# XLSX Tests
# ---------------------------------------------------------------------------

class TestExportXlsx:
    def test_creates_file(self, tmp_path, monkeypatch):
        sd = _create_mock_landscape(tmp_path)
        output = os.path.join(sd, "test-landscape.xlsx")

        sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "..", ".."))
        from cli_anything.cortellis.skills.landscape.recipes.export_xlsx import export_xlsx

        result = export_xlsx(sd, "Test Indication", output)
        assert os.path.exists(result)
        assert result.endswith(".xlsx")

    def test_sheet_names(self, tmp_path):
        sd = _create_mock_landscape(tmp_path)
        output = os.path.join(sd, "test-landscape.xlsx")

        from cli_anything.cortellis.skills.landscape.recipes.export_xlsx import export_xlsx
        export_xlsx(sd, "Test Indication", output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        names = wb.sheetnames
        assert "Pipeline by Phase" in names
        assert "Company Rankings" in names
        assert "Mechanism Analysis" in names
        assert "Deals" in names
        assert "Opportunity Matrix" in names
        wb.close()

    def test_pipeline_sheet_has_rows(self, tmp_path):
        sd = _create_mock_landscape(tmp_path)
        output = os.path.join(sd, "test-landscape.xlsx")

        from cli_anything.cortellis.skills.landscape.recipes.export_xlsx import export_xlsx
        export_xlsx(sd, "Test Indication", output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["Pipeline by Phase"]
        # Header + 5 data rows (2 launched + 1 phase3 + 2 phase2)
        assert ws.max_row >= 6
        wb.close()

    def test_company_sheet_has_cpi(self, tmp_path):
        sd = _create_mock_landscape(tmp_path)
        output = os.path.join(sd, "test-landscape.xlsx")

        from cli_anything.cortellis.skills.landscape.recipes.export_xlsx import export_xlsx
        export_xlsx(sd, "Test Indication", output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["Company Rankings"]
        # Check that CPI values are present — search for 85 as float or string
        found_cpi = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell in row:
                if cell is not None and ("85" in str(cell) or cell == 85.0):
                    found_cpi = True
                    break
        assert found_cpi, f"CPI 85.0 not found. Row 2 values: {[c for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]}"
        wb.close()

    def test_deals_without_financials(self, tmp_path):
        """Deals sheet works without deal_financials.csv."""
        sd = _create_mock_landscape(tmp_path)
        output = os.path.join(sd, "test-landscape.xlsx")

        from cli_anything.cortellis.skills.landscape.recipes.export_xlsx import export_xlsx
        export_xlsx(sd, "Test Indication", output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["Deals"]
        assert ws.max_row >= 3  # header + 2 deals
        wb.close()

    def test_deals_with_financials(self, tmp_path):
        """Deals sheet includes financial columns when deal_financials.csv exists."""
        sd = _create_mock_landscape(tmp_path)
        # Add deal_financials.csv
        _write_csv(os.path.join(sd, "deal_financials.csv"),
            ["deal_id", "title", "principal", "partner", "type", "date",
             "upfront_payment", "milestone_payments", "royalty_rate",
             "total_deal_value", "deal_currency", "financial_terms_text"],
            [
                ["D001", "Deal 1", "Company A", "Company B", "Licensing", "2026-03-01",
                 "$100M", "$1.5B", "15%", "$1.6B", "USD", "Upfront plus milestones"],
            ])
        output = os.path.join(sd, "test-landscape.xlsx")

        from cli_anything.cortellis.skills.landscape.recipes.export_xlsx import export_xlsx
        export_xlsx(sd, "Test Indication", output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["Deals"]
        # Should have more columns than without financials
        assert ws.max_column >= 8
        wb.close()


# ---------------------------------------------------------------------------
# PPTX Tests
# ---------------------------------------------------------------------------

class TestExportPptx:
    def test_creates_file(self, tmp_path):
        sd = _create_mock_landscape(tmp_path)
        output = os.path.join(sd, "test-landscape.pptx")

        from cli_anything.cortellis.skills.landscape.recipes.export_pptx import export_pptx
        result = export_pptx(sd, "Test Indication", output)
        assert os.path.exists(result)
        assert result.endswith(".pptx")

    def test_slide_count(self, tmp_path):
        sd = _create_mock_landscape(tmp_path)
        output = os.path.join(sd, "test-landscape.pptx")

        from cli_anything.cortellis.skills.landscape.recipes.export_pptx import export_pptx
        export_pptx(sd, "Test Indication", output)

        from pptx import Presentation
        prs = Presentation(output)
        # Should have at least 5 slides (title + summary + pipeline + competitive + sources)
        assert len(prs.slides) >= 5

    def test_empty_landscape(self, tmp_path):
        """Export with empty landscape directory still produces a file."""
        sd = str(tmp_path / "raw" / "empty")
        os.makedirs(sd)
        output = os.path.join(sd, "test-landscape.pptx")

        from cli_anything.cortellis.skills.landscape.recipes.export_pptx import export_pptx
        result = export_pptx(sd, "Empty Test", output)
        assert os.path.exists(result)


# ---------------------------------------------------------------------------
# Additional XLSX Tests
# ---------------------------------------------------------------------------


class TestExportXlsxAutoFilter:
    def test_pipeline_sheet_has_auto_filter(self, tmp_path):
        """Verify auto-filter is set on Pipeline by Phase sheet."""
        sd = _create_mock_landscape(tmp_path)
        output = os.path.join(sd, "test.xlsx")
        from cli_anything.cortellis.skills.landscape.recipes.export_xlsx import export_xlsx
        export_xlsx(sd, "Test", output)
        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["Pipeline by Phase"]
        assert ws.auto_filter.ref is not None
        wb.close()


class TestExportXlsxEmptyLandscape:
    def test_empty_dir_produces_file(self, tmp_path):
        """Export with empty landscape directory still produces a file."""
        sd = str(tmp_path / "raw" / "empty")
        os.makedirs(sd)
        output = os.path.join(sd, "test.xlsx")
        from cli_anything.cortellis.skills.landscape.recipes.export_xlsx import export_xlsx
        result = export_xlsx(sd, "Empty", output)
        assert os.path.exists(result)
